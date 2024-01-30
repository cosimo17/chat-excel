import fontTools.misc.cython
from openpyxl import load_workbook, workbook
from openpyxl.cell.cell import MergedCell
from src.observer import Observer
from src.memo import ModificationType
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image
import os


class ExcelAgent(Observer):
    def __init__(self):
        self.wb = None
        self.ws = None
        self.answer_ws = None
        self.plot_ws = None
        self.num_rows = 0
        self.num_cols = 0

    def load(self, filename):
        wb = load_workbook(filename=filename)
        if self.is_valid(wb):
            self.wb = wb
            self.ws = wb.active
            self.num_rows = self.ws.max_row
            self.num_cols = self.ws.max_column
            self.answer_ws = wb.create_sheet("AI answer")
            return True
        return False

    def save(self, filename, fig_dir=None):
        if not self.is_opened():
            return
        if fig_dir is not None:
            self.insert_image(fig_dir)
        self.wb.save(filename)

    def is_valid(self, wb):
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    return False
        return True

    def is_opened(self):
        return self.wb is not None and self.ws is not None

    def _update(self, subject):
        if not self.is_opened():
            return
        modifications = subject.modification
        mtype = modifications.mtype
        item_infos = modifications.item_infos
        if len(item_infos) == 0:
            return
        if mtype == ModificationType.NEW_TABLE:
            df = modifications.df
            columns = list(list(df.columns))
            # reset the answer sheet
            self.answer_ws = None
            del self.wb['AI answer']
            self.answer_ws = self.wb.create_sheet("AI answer")
            column_names = {}
            for item_info in item_infos:
                index = item_info.index
                text = item_info.text
                dtype = item_info.dtype
                data = dtype(text)
                col_name = item_info.column_name
                column_names[index[1]] = col_name
                excel_cell_index = self.index_to_excel_index(*index, row_bias=1)
                self.answer_ws[excel_cell_index] = data
                style = self.translate_style(item_info)
                self.apply_style(self.answer_ws, excel_cell_index, style)
            for k, v in column_names.items():
                index = (0, k)
                excel_cell_index = self.index_to_excel_index(*index)
                self.answer_ws[excel_cell_index] = v
        if mtype == ModificationType.UPDATE_INPLACE:
            for item_info in item_infos:
                index = item_info.index
                text = item_info.text
                dtype = item_info.dtype
                try:
                    data = dtype(text)
                except:
                    data = float(text)
                excel_cell_index = self.index_to_excel_index(*index)
                self.ws[excel_cell_index] = data
                style = self.translate_style(item_info)
                self.apply_style(self.ws, excel_cell_index, style)

    def insert_row(self, item_infos):
        for item_info in item_infos:
            index = item_info.index
            text = item_info.text
            dtype = item_info.dtype
            data = dtype(text)
            excel_cell_index = self.index_to_excel_index(*index)
            self.ws[excel_cell_index] = data
            style = self.translate_style(item_info)
            self.apply_style(self.ws, excel_cell_index, style)

    def insert_column(self, item_infos):
        for item_info in item_infos:
            index = item_info.index
            text = item_info.text
            dtype = item_info.dtype
            data = dtype(text)
            excel_cell_index = self.index_to_excel_index(*index)
            self.ws[excel_cell_index] = data
            style = self.translate_style(item_info)
            self.apply_style(self.ws, excel_cell_index, style)

    def insert_scalar(self, item_infos):
        item_info = item_infos[0]
        index = item_info.index
        text = item_info.text
        dtype = item_info.dtype
        data = dtype(text)
        excel_cell_index = self.index_to_excel_index(*index)
        self.ws[excel_cell_index] = data
        style = self.translate_style(item_info)
        self.apply_style(self.ws, excel_cell_index, style)

    def delete_scalar(self, item_infos):
        item_info = item_infos[0]
        row, column = item_info.index

        self.ws.cell(row=row, column=column, value=None)

    def delete_row(self, item_infos):
        item_info = item_infos[0]
        row, column = item_info.index
        self.ws.delete_rows(row)

    def delete_column(self, item_infos):
        item_info = item_infos[0]
        row, column = item_info.index
        for row_cells in self.ws.iter_rows():
            del row_cells[column]

    def update_inplace(self, item_infos):
        item_info = item_infos[0]
        index = item_info.index
        text = item_info.text
        dtype = item_info.dtype
        data = dtype(text)
        excel_cell_index = self.index_to_excel_index(*index)
        self.ws[excel_cell_index] = data
        style = self.translate_style(item_info)
        self.apply_style(self.ws, excel_cell_index, style)

    def index_to_excel_index(self, row, col, row_bias=0):
        excel_col = ''
        if col == 0:
            excel_col = 'A'
        else:
            col += 1
            while col > 0:
                remainder = (col - 1) % 26
                result = chr(ord('A') + remainder)
                excel_col = result + excel_col
                col = (col - 1) // 26

        excel_row = row + 1 + row_bias
        return f'{excel_col}{excel_row}'

    def translate_style(self, item_info):
        font = item_info.font
        font_size = item_info.font_size
        font_color = item_info.font_color.name()
        bg_color = item_info.bg_color.name()
        font_name = font.family()
        return font_name, font_size, 'FF' + font_color[1:], 'FF' + bg_color[1:]

    def apply_style(self, sheet, index, style):
        font_name, font_size, font_color, bg_color = style
        font = Font(name=font_name,
                    size=font_size,
                    color=font_color)
        fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color=bg_color)
        sheet[index].font = font
        sheet[index].fill = fill

    def insert_image(self, fig_dir):
        imgnames = os.listdir(fig_dir)
        if len(imgnames) == 0:
            return
        for i, imgname in enumerate(imgnames):
            plot_ws = self.wb.create_sheet("plot{}".format(i+1))
            imgname = os.path.join(fig_dir, imgname)
            img = Image(imgname)
            plot_ws.add_image(img, 'A{}'.format(i + 1))
