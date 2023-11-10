from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QTableWidget, \
    QTableWidgetItem, QHBoxLayout, QTextEdit, QShortcut, QMenuBar, \
    QMenu, QAction, QFileDialog, QScrollArea, QPushButton, QLabel, QTabWidget, QComboBox
from PyQt5.QtGui import QKeySequence, QIcon
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal
import sys
import openpyxl
from richtext_display import CodeEditor
from plotwin import PlotWidget
from memo import TableMemo, InplaceModification, Modification
import pandas as pd
from copy import deepcopy
from utis import withoutconnect, wrap_code, decodestdoutput, extract_func_info
from utis import is_assignment_statement
from interpreter import PythonInterpreter
from chatgpt import ChatBot
from prompt_template import prompt, chart_prompt
from openai.error import APIError
from enum import Enum


class QChatBot(QThread):
    res_signal = pyqtSignal(tuple)

    def __init__(self, prompts, default_answer, exception_answer):
        super(QChatBot, self).__init__()
        self.formatted_prompt = prompts
        self.bot = ChatBot()
        self.default_answer = default_answer
        self.exception_answer = exception_answer

    def run(self):
        if self.formatted_prompt == '':
            answer = self.default_answer
        else:
            try:
                answer, token_count = self.bot.get_response(self.formatted_prompt)
                answer = '\n#A:\n' + answer + '\n\n'
            except APIError:
                answer, token_count = self.exception_answer, 0
        self.res_signal.emit((answer, token_count))


class QInterpreter(QThread):
    res_signal = pyqtSignal(tuple)

    def __init__(self, code):
        super(QInterpreter, self).__init__()
        self.code = code
        self.interpreter = PythonInterpreter()

    def run(self):
        response = self.interpreter.execute(self.code)
        self.res_signal.emit(response)


class ChatWidget(QWidget):
    def __init__(self, main_win):
        super(ChatWidget, self).__init__()
        self.vbox = QVBoxLayout()

        # readonly richtext editer to display the chat history
        self.chat_history = CodeEditor()
        self.chat_history.setReadOnly(True)
        self.chat_history.setLineWrapMode(QTextEdit.NoWrap)
        # add scroll area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidget(self.chat_history)
        self.chat_history.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.chat_history.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.vbox.addWidget(self.chat_history, 4)

        self.switch_mode_box = QComboBox()
        self.switch_mode_box.addItem("Chat")
        self.switch_mode_box.addItem("Plot")
        self.switch_mode_box.currentIndexChanged.connect(main_win.switch_mode)
        self.mode_info = QLabel("mode: ")
        hbox2 = QHBoxLayout()
        # hbox2.addSpacing(20)
        hbox2.addWidget(self.mode_info)
        hbox2.addWidget(self.switch_mode_box)
        hbox2.addStretch(1)
        self.vbox.addLayout(hbox2)
        self.user_input = QTextEdit()
        self.vbox.addWidget(self.user_input, 1)

        self.token_logger = QLabel("Token:0   Cost:$0.0")
        self.vbox.addWidget(self.token_logger)
        self.setLayout(self.vbox)

    def set_token_usage(self, token):
        cost = token / 1000 * 0.002
        text = "Token: {} Cost: ${:.4f}".format(token, cost)
        self.token_logger.setText(text)


class Mode(Enum):
    CHAT_MODE = 1
    PLOT_MODE = 2


class Main(QWidget):
    def __init__(self):
        super(Main, self).__init__()
        self.setWindowTitle("chat-excel")
        self.loaded = False
        self.collapsed = False
        self.chat_widgets = None
        self.hbox = self.vbox = None
        self.table_widget = None
        self.dataframe = None
        self.code = None
        self.chat_thread = None
        self.interpreter_thread = None
        self.stdout = None
        self.stderror = None
        self.mode = Mode.CHAT_MODE
        self.token_count = 0
        self.default_answer = '\n#A:\n' + '请先打开需要处理的excel!\n\n'
        self.exception_answer = '\n#A:\n' + "抱歉，AI助理响应失败，请稍后再试" + '\n\n'
        self.recoder = TableMemo(self)
        self.init_ui()
        self.register_shortcut()
        self.init_table()
        self.bot = ChatBot()

    def init_ui(self):
        hbox = QHBoxLayout()
        vbox = QVBoxLayout()
        hbox.setContentsMargins(0, 0, 0, 0)
        self.hbox = hbox
        self.vbox = vbox
        # chat_widget = QWidget()
        # container to display excel data
        self.table_widget = QTableWidget()
        self.table_widget.itemChanged.connect(self.handle_item_changed)
        hbox.addWidget(self.table_widget, 3)

        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.West)
        self.tabs.setMovable(True)

        self.chat_widget = ChatWidget(self)
        self.plot_widget = PlotWidget()
        self.tabs.addTab(self.chat_widget, "chat")
        self.tabs.addTab(self.plot_widget, 'plot')
        hbox.addWidget(self.tabs, 2)
        # hbox.addWidget(self.chat_widget, 2)

        # menu setting
        menubar = QMenuBar()
        fileMenu = QMenu("&File", self)
        openAct = QAction('Open', self)
        openAct.triggered.connect(self.open_excel)
        fileMenu.addAction(openAct)
        saveAvt = QAction('Save', self)
        saveAvt.triggered.connect(self.file_save)
        fileMenu.addAction(saveAvt)
        menubar.addMenu(fileMenu)
        # Creating menus using a title
        editMenu = menubar.addMenu("&Edit")
        undoAct = QAction('Undo', self)
        undoAct.triggered.connect(self.undo_modification)
        editMenu.addAction(undoAct)
        helpMenu = menubar.addMenu("&Help")
        hbox.setMenuBar(menubar)

        self.setLayout(hbox)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.type_one_by_one)
        self.current_index = 0

    def register_shortcut(self):
        # hotkey
        # send chat message
        self.shortcut = QShortcut(QKeySequence(Qt.CTRL + Qt.Key_Q), self)
        self.shortcut.activated.connect(self.chat)

        # hide chat widget
        self.shortcut = QShortcut(QKeySequence(Qt.CTRL + Qt.Key_A), self)
        self.shortcut.activated.connect(self.collapse_chat_widget)

        # undo modification
        self.shortcut = QShortcut(QKeySequence(Qt.CTRL + Qt.Key_Z), self)
        self.shortcut.activated.connect(self.undo_modification)

    @withoutconnect
    def init_table(self):
        self.table_widget.setRowCount(20)
        self.table_widget.setColumnCount(20)
        for row in range(20):
            for col in range(20):
                self.table_widget.setItem(row, col, QTableWidgetItem(''))

    @withoutconnect
    def open_excel(self):
        filename, filetype = QFileDialog.getOpenFileName(self, "select excel file", "", "*.xlsx;;*.xls;;All Files(*)")
        if not filename:
            return
        self.load_data_with_pandas(filename)
        self.loaded = True

    def load_data(self, excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        self.table_widget.setRowCount(sheet.max_row)
        self.table_widget.setColumnCount(sheet.max_column)

        list_values = list(sheet.values)
        self.table_widget.setHorizontalHeaderLabels(list_values[0])

        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.table_widget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1

    def load_data_with_pandas(self, excel_file):
        df = pd.read_excel(excel_file)
        shape = df.shape
        rows, cols = shape
        self.table_widget.setRowCount(rows + 1)
        self.table_widget.setColumnCount(cols)
        self.table_widget.setHorizontalHeaderLabels(list(df.columns))
        for i in range(rows):
            for j in range(cols):
                value = df.iloc[i, j]
                self.table_widget.setItem(i, j, QTableWidgetItem(str(value)))
        for j in range(cols):
            self.table_widget.setItem(rows, j, QTableWidgetItem(''))
        self.dataframe = deepcopy(df)

    def save(self):
        return self.modification

    def restore(self, modification):
        if modification is None:
            return
        if isinstance(modification, InplaceModification):
            rows = modification.row_indexs
            cols = modification.col_indexs
            values = modification.values
            for i in range(len(modification)):
                row, col, value = rows[i], cols[i], values[i]
                if value != 'chatgpt-placeholder':
                    self.table_widget.setItem(row, col, QTableWidgetItem(str(value)))
                    d = self.dataframe.iloc[row, col]
                    self.dataframe.iloc[row, col] = type(d)(value)
                else:
                    self.table_widget.setItem(row, col, QTableWidgetItem(''))
        elif isinstance(modification, Modification):
            df = modification.df
            shape = df.shape
            row_count = self.table_widget.rowCount()
            col_count = self.table_widget.columnCount()
            if shape == self.dataframe.shape:  # Update the entire table in place
                for i in range(shape[0]):
                    for j in range(shape[1]):
                        value = df.iloc[i, j]
                        self.table_widget.setItem(i, j, QTableWidgetItem(str(value)))
            else:
                # delete column
                if shape[0] == self.dataframe.shape[0]:
                    self.table_widget.setColumnCount(col_count - 1)

                # delete row
                if shape[1] == self.dataframe.shape[1]:
                    self.table_widget.setRowCount(row_count - 1)

    @withoutconnect
    def undo_modification(self):
        self.recoder.undo()

    def save_checkpoint(self):
        self.recoder.backup()

    def handle_item_changed(self, item):
        if self.dataframe is None:
            return
        row = item.row()
        col = item.column()
        value = item.text()
        self.modification = InplaceModification([row], [col], [self.dataframe.iloc[row, col]])
        d = self.dataframe.iloc[row, col]
        self.dataframe.iloc[row, col] = type(d)(value)
        self.save_checkpoint()

    def collapse_chat_widget(self):
        if not self.collapsed:
            widget = self.hbox.itemAt(1).widget()
            widget.setVisible(False)
            # backup chat widget
            self.chat_widget = widget
            self.collapsed = True
        else:
            self.chat_widget.setVisible(True)
            self.collapsed = False

    def type_one_by_one(self):
        if not self.answer:
            return
        if self.current_index < len(self.answer):
            next_char = self.answer[self.current_index]
            self.chat_widget.chat_history.insertPlainText(next_char)
            self.current_index += 1
        else:
            self.code = self.answer
            self.timer.stop()
            self.answer = None
            if not (self.code == self.default_answer or self.code == self.exception_answer):
                # execute the code and display the result
                self.execute()
                # self.switch_button()

    def execute(self):
        if self.mode == Mode.CHAT_MODE:
            code = wrap_code(self.code, self.dataframe)
            self.interpreter_thread = QInterpreter(code)
            self.interpreter_thread.res_signal.connect(self.receive_output)
            self.interpreter_thread.start()
        elif self.mode == Mode.PLOT_MODE:
            df = self.dataframe
            func_names, func_args, func_kwargs = extract_func_info(self.code, df)
            self.plot_widget.new_axes()
            self.plot_widget.call_func(func_names, func_args, func_kwargs)
            self.plot_widget.add_figure()
            self.tabs.setCurrentIndex(1)

    @withoutconnect
    def insert_result(self):
        if self.stdout is None:
            return
        row_count = self.table_widget.rowCount()
        col_count = self.table_widget.columnCount()
        # result is DataFrame
        if isinstance(self.stdout, pd.DataFrame):
            df = self.stdout
            shape = df.shape
            if shape == self.dataframe.shape:  # Update the entire table in place
                for i in range(shape[0]):
                    for j in range(shape[1]):
                        value = df.iloc[i, j]
                        self.table_widget.setItem(i, j, QTableWidgetItem(str(value)))
            else:
                if shape[0] == self.dataframe.shape[0]:  # insert new column
                    self.table_widget.setColumnCount(col_count + 1)
                    for i in range(shape[0]):
                        value = df.iloc[i, shape[1] - 1]
                        self.table_widget.setItem(i, col_count, QTableWidgetItem(str(value)))
                if shape[1] == self.dataframe.shape[1]:  # insert new row
                    self.table_widget.setRowCount(row_count + 1)
                    for j in range(shape[1]):
                        value = df.iloc[shape[0] - 1, j]
                        self.table_widget.setItem(row_count - 1, j, QTableWidgetItem(str(value)))
            self.modification = Modification(deepcopy(self.dataframe))
            self.save_checkpoint()
            self.dataframe = df

        # result is a scalar
        else:
            self.table_widget.setItem(row_count - 1, 0, QTableWidgetItem(str(self.stdout)))
            self.modification = InplaceModification([row_count - 1], [0], ['chatgpt-placeholder'])
            self.save_checkpoint()
        self.stdout = None

    def receive_output(self, output):
        stdout, stderror = output
        try:
            stdout = decodestdoutput(stdout)
        except Exception:
            return
        self.stdout = stdout
        self.stderro = stderror
        # display the output and handle the error
        self.insert_result()

    def chat(self):
        self.table_widget.maximumWidth()
        if self.chat_widget.user_input.hasFocus():
            message = self.chat_widget.user_input.toPlainText().strip()
            task = message
            if message:
                message = "Q:\n{}".format(message)
                self.chat_widget.chat_history.append(message)
                self.chat_widget.user_input.clear()
                formatted_prompt = self.format_prompt(task)
                # get response from llm
                # using QThread to avoid GUI freeze
                self.chat_thread = QChatBot(formatted_prompt, self.default_answer, self.exception_answer)
                self.chat_thread.res_signal.connect(self.receive_answer)
                self.chat_thread.start()

        return

    def format_prompt(self, task):
        template_prompt = prompt if self.mode == Mode.CHAT_MODE else chart_prompt
        formatted_prompt = '' if self.dataframe is None \
            else template_prompt.format(self.dataframe.shape,
                                        self.dataframe.head(3),
                                        self.dataframe.dtypes, task)
        return formatted_prompt

    def receive_answer(self, res):
        answer, token_count = res
        self.answer = answer
        self.token_count += token_count
        self.chat_widget.set_token_usage(self.token_count)
        self.current_index = 0
        self.timer.start(30)

    def file_save(self):
        file_filter = "*.xlsx;;*.xls;;All Files(*)"
        path, _ = QFileDialog.getSaveFileName(self, "Save excel file", "", file_filter)

        if not path:
            return
        self.dataframe.to_excel(path)

    def switch_mode(self, index):
        self.mode = Mode.PLOT_MODE if self.chat_widget.switch_mode_box.currentText() == "Plot" else Mode.CHAT_MODE


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Main()
    window.showMaximized()
    app.exec_()
